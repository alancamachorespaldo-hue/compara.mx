"""
excel_a_codigo.py — Convierte Comparalo.mx (2).xlsx en código JS para index.html
==================================================================================
USO: python excel_a_codigo.py
Genera: codigo_generado.txt — copia y pega en index.html
"""

import re
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

ARCHIVO = r"C:\Users\alanc\Downloads\GestorCompara\Comparalo.mx (2).xlsx"
SALIDA  = r"C:\Users\alanc\Downloads\GestorCompara\codigo_generado.txt"

# ── Siguiente ID disponible por sección (ajusta si ya tienes IDs más altos)
NEXT_ID = {
    "elec":    20,   # IDs 1-13 usados
    "ciudad":  500,  # nueva sección
}

def extraer_precio(texto):
    """Extrae el precio de venta (descartando cantidades MSI mensuales)."""
    # Cortar texto antes de "meses sin intereses" para ignorar montos MSI
    texto_limpio = re.split(r'\d+\s*meses?\s*sin\s*intereses', texto, flags=re.IGNORECASE)[0]
    # Buscar números que siguen a un $ en línea propia o inline
    candidatos = re.findall(r'^\$?\s*\n?([\d]{1,3}(?:,\d{3})+)\s*$', texto_limpio, re.MULTILINE)
    if not candidatos:
        candidatos = re.findall(r'\$([\d]{1,3}(?:,\d{3})+)', texto_limpio)
    if not candidatos:
        candidatos = re.findall(r'(?<!\d)([\d]{4,6})(?!\d)', texto_limpio)
    nums = []
    for p in candidatos:
        try:
            nums.append(int(str(p).replace(',', '')))
        except Exception:
            pass
    # Filtrar: precio real >= 500, descartar centavos/decimales
    significativos = sorted([x for x in nums if x >= 500])
    if not significativos:
        return 0
    # Si hay dos precios (original y descuento), tomar el más bajo (precio final)
    return significativos[0] if len(significativos) == 1 else significativos[-1] if significativos[-1] < 2 * significativos[0] else significativos[0]

def extraer_stars_reviews(texto):
    stars, reviews = 0, 0
    m = re.search(r'(\d+\.\d+)\s*\n?Calificación', texto)
    if not m:
        m = re.search(r'^(\d\.\d)$', texto, re.MULTILINE)
    if m:
        stars = float(m.group(1))
    r = re.search(r'\((\d+)\)', texto)
    if r:
        reviews = int(r.group(1))
    return stars, reviews

def extraer_nombre(texto):
    """Primera línea no vacía = nombre del producto."""
    for linea in texto.split('\n'):
        linea = linea.strip()
        if not linea or linea.startswith('$') or linea.startswith('('):
            continue
        # Quitar sufijos de calificación embebidos
        linea = re.sub(r'\s*★[\d.]+.*$', '', linea)
        linea = re.sub(r'\s*\|\s*(Amazon|Mercado Libre).*$', '', linea, flags=re.IGNORECASE)
        linea = re.sub(r'\s*\d+\.\d+\s*$', '', linea).strip()
        # Cortar nombres muy largos en pipe o descripción larga
        if ' | ' in linea:
            linea = linea.split(' | ')[0].strip()
        return linea[:120]
    return ""

def extraer_specs(desc):
    """Extrae specs de la columna Descripción."""
    specs = {}
    if not desc:
        return specs
    patrones = {
        'rueda':  r'Rodado[:\s]+(\d+[cC]?)',
        'vel':    r'(?:Cantidad de velocidades|velocidades)[:\s]+(\d+)',
        'peso_kg':r'Peso máximo soportado[:\s]+(\d+)',
        'peso':   r'^Peso\s*\n(\d+(?:\.\d+)?\s*kg)',
    }
    for k, pat in patrones.items():
        m = re.search(pat, desc, re.IGNORECASE | re.MULTILINE)
        if m:
            specs[k] = m.group(1).strip()
    return specs

def es_mv(texto):
    return 'MÁS VENDIDO' in texto.upper() or 'MAS VENDIDO' in texto.upper()

def procesar_hoja(ws, sec_nombre, start_row=2):
    items = []
    vistos = set()
    for r in range(start_row, ws.max_row + 1):
        link = ws.cell(r, 1).value
        img  = ws.cell(r, 2).value
        txt  = ws.cell(r, 3).value
        desc = ws.cell(r, 4).value

        if not link and not txt:
            continue
        if not txt:
            continue

        link = str(link).strip() if link else ''
        img  = str(img).strip()  if img  else ''
        txt  = str(txt).strip()

        nombre  = extraer_nombre(txt)
        precio  = extraer_precio(txt)
        stars, reviews = extraer_stars_reviews(txt)
        specs   = extraer_specs(str(desc) if desc else '')
        mv      = es_mv(txt)

        if not nombre:
            continue
        if link and link in vistos:
            continue
        if link:
            vistos.add(link)

        items.append({
            'nombre':  nombre,
            'precio':  precio,
            'link':    link,
            'img':     img,
            'stars':   stars,
            'reviews': reviews,
            'rueda':   specs.get('rueda', ''),
            'vel':     specs.get('vel', ''),
            'kg':      specs.get('peso_kg', ''),
            'mv':      mv,
        })
    return items

def generar_objeto(item, id_num, tipo='mtb'):
    def esc(s): return str(s).replace("'", "\\'")
    link = item['link']
    es_amz = 'amazon.com' in link.lower()
    nombre = esc(item['nombre'])
    marca  = esc(item['nombre'].split()[0].capitalize())
    precio = item['precio'] or 0
    precio_ml = 0 if es_amz else precio
    precio_amz = precio if es_amz else 0
    rueda = item['rueda']
    if rueda and not rueda.endswith('"'):
        rueda = rueda + '"'
    vel = item['vel'] or '—'
    stars = item['stars'] or 0
    reviews = item['reviews'] or 0
    carga = item['kg'] or 0

    # Badges
    bdg = []
    if es_amz:
        bdg.append("{t:'AMZ',c:'b-amz'}")
    else:
        bdg.append("{t:'ML',c:'b-ml'}")
    if item['mv']:
        bdg.append("{t:'Más vendida',c:'b-best'}")
    if reviews >= 100:
        bdg.append(f"{{t:'{reviews:,} reseñas',c:'b-pop'}}")
    badges_str = '[' + ','.join(bdg) + ']'

    link_amz = f"linkAmz:'{esc(link)}'," if es_amz else "linkAmz:'',"
    link_ml  = "linkML:''," if es_amz else f"linkML:'{esc(link)}',"

    lines = [
        f"  {{id:{id_num},nombre:'{nombre}',marca:'{marca}',tipo:'{tipo}',",
        f"   precio:{precio},velocidad:0,peso:0,rueda:'{rueda}',velocidades:'{vel}',cargaMax:{carga},",
        f"   estrellas:{stars},resenas:{reviews},",
        f"   badges:{badges_str},",
        f"   {link_amz}",
        f"   {link_ml}",
        f"   img:'{esc(item['img'])}'}},",
    ]
    return '\n'.join(lines)

def main():
    print("=" * 60)
    print("  Excel → Código JS — comparalo.mx")
    print("=" * 60)

    wb = openpyxl.load_workbook(ARCHIVO, data_only=True)
    print(f"Hojas encontradas: {wb.sheetnames}\n")

    output = []
    total = 0

    # Hoja Catálogo Bicis (eléctricas) — headers en fila 2, datos desde fila 3
    if 'Catálogo Bicis' in wb.sheetnames:
        ws = wb['Catálogo Bicis']
        items = procesar_hoja(ws, 'elec', start_row=3)
        if items:
            output.append("// ══ ELECTRICAS — pegar en const ELECTRICAS = [ ... ]")
            id_n = NEXT_ID['elec']
            for it in items:
                output.append(generar_objeto(it, id_n, tipo='plegable'))
                id_n += 1
            total += len(items)
            print(f"Catálogo Bicis (eléctricas): {len(items)} productos")

    # Clasificador por palabras clave para Bicicletas Ciudad
    def clasificar(nombre):
        n = nombre.lower()
        if any(w in n for w in ['gravel tucson','tucson','sora','apex']):
            return 'gravel'
        if any(w in n for w in ['gravel','fixie','r700','700c','ruta','road','technik','strada','benotto','alubike','monk','armstrong','ram road','atenea']):
            return 'ruta'
        return 'mtb'

    if 'Bicicletas Ciudad' in wb.sheetnames:
        ws = wb['Bicicletas Ciudad']
        items = procesar_hoja(ws, 'ciudad', start_row=2)
        if items:
            montana = [it for it in items if clasificar(it['nombre']) == 'mtb']
            ruta    = [it for it in items if clasificar(it['nombre']) == 'ruta']
            gravel  = [it for it in items if clasificar(it['nombre']) == 'gravel']

            for label, grupo, sec in [
                ("// ══ MONTANA — pegar en const MONTANA = [ ... ]", montana, 'mtb'),
                ("\n// ══ RUTA — pegar en const RUTA = [ ... ]", ruta, 'ruta'),
                ("\n// ══ GRAVEL — pegar en const GRAVEL = [ ... ]", gravel, 'gravel'),
            ]:
                if grupo:
                    output.append(label)
                    id_n = NEXT_ID['ciudad']
                    for it in grupo:
                        output.append(generar_objeto(it, id_n, tipo=sec))
                        id_n += 1
                        NEXT_ID['ciudad'] += 1
            total += len(items)
            print(f"Bicicletas Ciudad: {len(items)} productos ({len(montana)} MTB, {len(ruta)} Ruta, {len(gravel)} Gravel)")

    with open(SALIDA, 'w', encoding='utf-8') as f:
        f.write('\n'.join(output))

    print(f"\nTotal: {total} productos")
    print(f"Guardado en: {SALIDA}")
    print("\nSiguiente paso:")
    print("  1. Abre codigo_generado.txt")
    print("  2. Revisa y ajusta los objetos")
    print("  3. Pega cada bloque en el array correcto de index.html")

if __name__ == "__main__":
    main()
    input("\nPresiona Enter para cerrar...")

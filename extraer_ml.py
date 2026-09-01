"""
extraer_ml.py — Extrae datos de productos de Mercado Libre por URL/ID
=====================================================================
USO:
  1. Pon los links o IDs de ML que quieras en la lista ITEMS abajo
  2. Ejecuta: python extraer_ml.py
  3. Revisa resultados_ml.xlsx y compártelo

Funciona sin necesidad de app certificada.
"""

import urllib.request
import urllib.parse
import urllib.error
import json
import ssl
import re
import sys
import time

# ─── CONFIGURACION ────────────────────────────────────────────────────────────
APP_ID     = "6900783320059393"
APP_SECRET = "FhsGJPFno3iP7AXJuaWTsAf11SKBqt43"

# Pega aqui los links de ML que quieras extraer:
ITEMS = [
    # ── Eléctricas ──────────────────────────────────────────
    "https://articulo.mercadolibre.com.mx/MLM-3720968014-bicicleta-electrica-uber-d460-plegable-750w-40kmh-2-pasaje-_JM",
    "https://meli.la/2cccAnq",
    "https://meli.la/2Eu51p1",
    "https://meli.la/14P2TRD",
    "https://meli.la/13mxVr4",
    "https://meli.la/28LXjta",
    "https://meli.la/2GGZZPS",
    "https://meli.la/1PbKNjX",
    # ── Montaña ─────────────────────────────────────────────
    "https://meli.la/1YqtNuV",
    "https://meli.la/16qSVsg",
    "https://meli.la/2GtxKLu",
    "https://meli.la/2Nu3SVW",
    # ── Convencionales / varias ──────────────────────────────
    "https://meli.la/1z195hd",
    "https://meli.la/2KKZNgU",
    "https://meli.la/1J6QUWi",
    "https://meli.la/2jyzTcR",
    "https://meli.la/1SHfv2C",
    "https://meli.la/1Ex8GXS",
    # ── Ruta ────────────────────────────────────────────────
    "https://meli.la/1noKL8E",
    "https://meli.la/17jmy7D",
    "https://meli.la/2DdcB9o",
    "https://meli.la/1sX2WZs",
    "https://meli.la/2T5fSoC",
    "https://meli.la/1n91RZ9",
    "https://meli.la/2riR7pt",
    "https://meli.la/2L1qxxF",
    "https://meli.la/2SXni1U",
    "https://meli.la/2reBECj",
    "https://meli.la/2DrvPBU",
    # ── Gravel ──────────────────────────────────────────────
    "https://meli.la/1wbKiie",
    "https://meli.la/2DEKrHB",
    "https://meli.la/2jcmFnH",
]
# ─────────────────────────────────────────────────────────────────────────────

_token = None
CTX = ssl.create_default_context()


def get_token():
    global _token
    if _token:
        return _token
    print("Obteniendo token...", end=" ", flush=True)
    body = urllib.parse.urlencode({
        "grant_type":    "client_credentials",
        "client_id":     APP_ID,
        "client_secret": APP_SECRET,
    }).encode()
    req = urllib.request.Request(
        "https://api.mercadolibre.com/oauth/token",
        data=body, method="POST",
        headers={"Content-Type": "application/x-www-form-urlencoded",
                 "Accept": "application/json"},
    )
    with urllib.request.urlopen(req, context=CTX, timeout=15) as r:
        _token = json.loads(r.read())["access_token"]
    print("OK")
    return _token


def fetch_page(url):
    """Obtiene el HTML de una página de producto ML."""
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor())
    opener.addheaders = [
        ("User-Agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0 Safari/537.36"),
        ("Accept", "text/html,application/xhtml+xml,*/*;q=0.9"),
        ("Accept-Language", "es-MX,es;q=0.9"),
    ]
    try:
        with opener.open(url, timeout=20) as r:
            return r.read().decode("utf-8", errors="ignore")
    except Exception as e:
        print(f"  [warn] fetch {url[:60]}: {e}")
        return ""


def parse_producto(html, item_id):
    """Extrae datos del JSON embebido en la página de ML."""
    # Buscar JSON-LD (schema.org)
    m = re.search(r'<script type="application/ld\+json">(.*?)</script>', html, re.DOTALL)
    ld = {}
    if m:
        try:
            ld = json.loads(m.group(1))
        except Exception:
            pass

    # Buscar precio en HTML
    precio = 0
    pm = re.search(r'"price"\s*:\s*([\d.]+)', html)
    if pm:
        try:
            precio = float(pm.group(1))
        except Exception:
            pass
    if not precio and ld.get("offers"):
        try:
            precio = float(ld["offers"].get("price", 0))
        except Exception:
            pass

    titulo = ld.get("name", "")
    if not titulo:
        tm = re.search(r'<h1[^>]*class="[^"]*ui-pdp-title[^"]*"[^>]*>(.*?)</h1>', html, re.DOTALL)
        if tm:
            titulo = re.sub(r'<[^>]+>', '', tm.group(1)).strip()

    # Imagen principal
    img = ""
    im = re.search(r'"https://http2\.mlstatic\.com/D_NQ_NP[^"]+?-O\.(jpg|webp)"', html)
    if im:
        img = im.group(0).strip('"')
    if not img:
        im2 = re.search(r'"https://http2\.mlstatic\.com/D_NQ_NP[^"]+?"', html)
        if im2:
            img = im2.group(0).strip('"')

    # Calificación
    stars, resenas = 0, 0
    rm = re.search(r'"ratingValue"\s*:\s*([\d.]+)', html)
    if rm:
        stars = round(float(rm.group(1)), 1)
    rc = re.search(r'"reviewCount"\s*:\s*(\d+)', html)
    if rc:
        resenas = int(rc.group(1))

    # Link limpio
    link = f"https://articulo.mercadolibre.com.mx/{item_id.replace('MLM', 'MLM-')}"

    return {
        "titulo": titulo[:80],
        "precio": precio,
        "img": img,
        "stars": stars,
        "resenas": resenas,
        "link": link,
        "item_id": item_id,
    }


def extraer_id(texto):
    """Extrae el MLM... de un link o devuelve el ID directo."""
    texto = texto.strip()
    if texto.upper().startswith("MLM"):
        return texto.upper()
    m = re.search(r"MLM-?(\d+)", texto, re.IGNORECASE)
    if m:
        return "MLM" + m.group(1)
    # meli.la o URLs cortas — seguir redirects con GET
    try:
        opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor())
        opener.addheaders = [
            ("User-Agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0 Safari/537.36"),
            ("Accept", "text/html,application/xhtml+xml,*/*"),
        ]
        with opener.open(texto, timeout=15) as r:
            final = r.url
            body = r.read(4000).decode("utf-8", errors="ignore")
        # buscar MLM en URL final
        m2 = re.search(r"MLM-?(\d+)", final, re.IGNORECASE)
        if m2:
            return "MLM" + m2.group(1)
        m3 = re.search(r"MLM-?(\d+)", body, re.IGNORECASE)
        if m3:
            return "MLM" + m3.group(1)
    except Exception as e:
        print(f"  [redirect error] {texto[:50]}: {e}")
    return None


def mejor_img(item):
    pics = item.get("pictures", [])
    if pics:
        url = pics[0].get("url", "")
        return url.replace("-O.jpg", "-O.jpg")  # ya es _O por defecto en /items
    return item.get("thumbnail", "").replace("-I.", "-O.")


def attr_val(attrs, attr_id):
    for a in attrs:
        if a.get("id") == attr_id:
            return a.get("value_name", "")
    return ""


def main():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: pip install openpyxl")
        sys.exit(1)

    if not ITEMS:
        print("\n" + "="*60)
        print("  No hay items. Edita la lista ITEMS en el script.")
        print("  Pega links de ML como:")
        print('  "https://www.mercadolibre.com.mx/bici-.../p/MLM..."')
        print("="*60)
        sys.exit(0)

    print("=" * 60)
    print("  Extractor de productos — Mercado Libre Mexico")
    print("=" * 60)

    ids = []
    for raw in ITEMS:
        item_id = extraer_id(raw)
        if item_id:
            ids.append(item_id)
        else:
            print(f"  [skip] No pude extraer ID de: {raw[:60]}")

    print(f"\n{len(ids)} items a procesar")

    COLS = [
        ("Publicar (SI/NO)", 18),
        ("Nombre",           52),
        ("Marca",            16),
        ("Precio MXN",       14),
        ("Link ML",          55),
        ("Imagen URL",       70),
        ("Estrellas",        12),
        ("Resenas",          10),
        ("Motor",            14),
        ("Autonomia km",     14),
        ("Velocidad kmh",    14),
        ("Rueda",            10),
        ("Velocidades",      13),
        ("Peso kg",          10),
        ("ID ML",            22),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    hf = PatternFill("solid", fgColor="006847")
    yf = PatternFill("solid", fgColor="FFF9C4")
    thin = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style="thin"),
        right=openpyxl.styles.Side(style="thin"),
        top=openpyxl.styles.Side(style="thin"),
        bottom=openpyxl.styles.Side(style="thin"),
    )

    for c, (h, w) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hf
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30

    row = 2
    for item_id in ids:
        print(f"  {item_id}...", end=" ", flush=True)
        page_url = f"https://articulo.mercadolibre.com.mx/{item_id.replace('MLM', 'MLM-')}"
        html = fetch_page(page_url)
        if not html:
            print("sin respuesta")
            continue

        p = parse_producto(html, item_id)
        if not p["titulo"] and not p["precio"]:
            print("no encontrado")
            continue

        ws.cell(row=row, column=1,  value="SI").fill = yf
        ws.cell(row=row, column=2,  value=p["titulo"])
        ws.cell(row=row, column=3,  value=p["titulo"].split()[0].capitalize())
        ws.cell(row=row, column=4,  value=p["precio"])
        ws.cell(row=row, column=5,  value=p["link"])
        ws.cell(row=row, column=6,  value=p["img"])
        ws.cell(row=row, column=7,  value=p["stars"] or "")
        ws.cell(row=row, column=8,  value=p["resenas"] or "")
        ws.cell(row=row, column=9,  value="")   # Motor — llenar manual si es eléctrica
        ws.cell(row=row, column=10, value="")   # Autonomia
        ws.cell(row=row, column=11, value="")   # Velocidad
        ws.cell(row=row, column=12, value="")   # Rueda
        ws.cell(row=row, column=13, value="")   # Velocidades
        ws.cell(row=row, column=14, value="")   # Peso
        ws.cell(row=row, column=15, value=item_id)
        for c in range(1, len(COLS)+1):
            ws.cell(row=row, column=c).border = thin

        lbl = f"${p['precio']:,.0f}" if p["precio"] else "sin precio"
        print(f"OK — {lbl} | {p['titulo'][:40]}")
        row += 1
        time.sleep(1)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    out = "resultados_ml.xlsx"
    wb.save(out)
    print(f"\nListo: {out} — {row-2} productos")


if __name__ == "__main__":
    main()
    input("\nPresiona Enter para cerrar...")

"""
buscar_ml.py — Buscador automatico de bicicletas en Mercado Libre Mexico
=========================================================================
REQUISITOS:
  1. Crea una app GRATIS en: https://developers.mercadolibre.com.mx/
     - Nombre: comparalo-scraper
     - Pega tu App ID y Secret Key abajo en CONFIGURACION
  2. Ejecuta: python buscar_ml.py

Genera: resultados_ml.xlsx — revisa y marca SI/NO, luego compartelo.
"""

import urllib.request
import urllib.parse
import urllib.error
import json
import ssl
import time
import sys

# ─── CONFIGURACION — pon tus credenciales aqui ────────────────────────────────
APP_ID     = "6900783320059393"
APP_SECRET = "FhsGJPFno3iP7AXJuaWTsAf11SKBqt43"
# ─────────────────────────────────────────────────────────────────────────────

BUSQUEDAS = [
    ("bicicleta electrica plegable", "elec"),
    ("bicicleta electrica urbana",   "elec"),
    ("bicicleta electrica montana",  "elec"),
]
LIMITE_POR_BUSQUEDA = 50
SITIO = "MLM"
# ─────────────────────────────────────────────────────────────────────────────

_token = None

def get_token():
    global _token
    if _token:
        return _token
    if not APP_ID or not APP_SECRET:
        print("\n" + "="*60)
        print("ERROR: Faltan las credenciales de ML.")
        print("Abre buscar_ml.py y coloca tu APP_ID y APP_SECRET.")
        print("Obtenlos en: https://developers.mercadolibre.com.mx/")
        print("="*60)
        sys.exit(1)

    print("Obteniendo token de ML...", end=" ", flush=True)
    body = urllib.parse.urlencode({
        "grant_type":    "client_credentials",
        "client_id":     APP_ID,
        "client_secret": APP_SECRET,
        "scope":         "read_catalog offline_access",
    }).encode()
    req = urllib.request.Request(
        "https://api.mercadolibre.com/oauth/token",
        data=body,
        method="POST",
        headers={"Content-Type": "application/x-www-form-urlencoded",
                 "Accept": "application/json"},
    )
    ctx = ssl.create_default_context()
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=15) as r:
            data = json.loads(r.read())
        _token = data["access_token"]
        print("OK")
        return _token
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print(f"\nError al obtener token: {e.code} {body}")
        sys.exit(1)


def api_get(path, auth=False):
    url = "https://api.mercadolibre.com" + path
    headers = {
        "Accept": "application/json",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0 Safari/537.36",
    }
    if auth:
        headers["Authorization"] = f"Bearer {get_token()}"
    req = urllib.request.Request(url, headers=headers)
    ctx = ssl.create_default_context()
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=15) as r:
            return json.loads(r.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        if e.code != 404:
            print(f"  [warn] {path}: HTTP {e.code}")
        return None
    except Exception as e:
        print(f"  [warn] {path}: {e}")
        return None


def buscar(query, seccion, limite=50):
    q = urllib.parse.quote(query)
    data = api_get(f"/sites/{SITIO}/search?q={q}&limit={limite}", auth=True)
    if data:
        results = data.get("results", [])
        for r in results:
            r["_seccion"] = seccion
        return results
    return []


def mejor_imagen(item):
    thumb = item.get("thumbnail", "")
    if thumb:
        # ML thumbnail URL: cambiar I.jpg -> O.jpg para mejor calidad
        thumb = thumb.replace("-I.jpg", "-O.jpg").replace("-I.webp", "-O.webp")
        thumb = thumb.replace("_I.", "_O.")
    return thumb


def limpiar_titulo(titulo):
    for sep in ["\n", " | ", " Agregar", " Visita", " Color "]:
        titulo = titulo.split(sep)[0]
    return titulo.strip()


def extraer_marca(titulo, attrs):
    # intentar desde atributos ML
    for a in attrs:
        if a.get("id") == "BRAND":
            return a.get("value_name", "")
    # fallback: primera palabra en mayusculas
    palabras = titulo.split()
    if palabras:
        return palabras[0].capitalize()
    return ""


def extraer_attr(attrs, attr_id):
    for a in attrs:
        if a.get("id") == attr_id:
            return a.get("value_name", "")
    return ""


def main():
    print("=" * 60)
    print("  Buscador de bicis — Mercado Libre Mexico")
    print("=" * 60)

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("\nERROR: Falta openpyxl. Instala con:  pip install openpyxl")
        sys.exit(1)

    todos = {}  # item_id -> item

    for query, seccion in BUSQUEDAS:
        print(f"\nBuscando: '{query}'...")
        items = buscar(query, seccion, LIMITE_POR_BUSQUEDA)
        nuevos = 0
        for it in items:
            if it["id"] not in todos:
                todos[it["id"]] = it
                nuevos += 1
        print(f"  {len(items)} resultados, {nuevos} nuevos")

    print(f"\nTotal unicos: {len(todos)} productos")
    print("Obteniendo resenas... (puede tardar 1-2 min)")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Electricas"

    COLS = [
        ("Publicar (SI/NO)",  18),
        ("Nombre",            52),
        ("Marca",             16),
        ("Precio MXN",        14),
        ("Link ML",           50),
        ("Imagen URL",        70),
        ("Estrellas",         12),
        ("Resenas",           10),
        ("Motor",             12),
        ("Autonomia km",      14),
        ("Velocidad kmh",     14),
        ("Rueda",             10),
        ("Velocidades",       13),
        ("Peso kg",           10),
        ("ID Mercado Libre",  22),
    ]

    head_fill = PatternFill("solid", fgColor="006847")
    pub_fill  = PatternFill("solid", fgColor="FFF9C4")
    no_fill   = PatternFill("solid", fgColor="FFE0E0")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for c, (h, w) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = head_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30

    row = 2
    for idx, (item_id, item) in enumerate(todos.items()):
        attrs = item.get("attributes", [])
        titulo = limpiar_titulo(item.get("title", ""))
        precio = item.get("price", 0)
        link   = item.get("permalink", "")
        marca  = extraer_marca(titulo, attrs)
        img    = mejor_imagen(item)

        motor     = extraer_attr(attrs, "ENGINE_TYPE") or extraer_attr(attrs, "MOTOR_POWER")
        autonomia = extraer_attr(attrs, "AUTONOMY") or extraer_attr(attrs, "MAX_RANGE")
        velocidad = extraer_attr(attrs, "MAX_SPEED")
        rueda     = extraer_attr(attrs, "WHEEL_SIZE")
        velocidades = extraer_attr(attrs, "GEARS_QUANTITY")
        peso      = extraer_attr(attrs, "WEIGHT")

        # Reviews — endpoint separado
        stars, resenas = 0, 0
        rev = api_get(f"/reviews/item/{item_id}", auth=True)
        if rev:
            stars   = round(rev.get("rating_average", 0), 1)
            resenas = rev.get("paging", {}).get("total", 0)

        ws.cell(row=row, column=1,  value="SI").fill = pub_fill
        ws.cell(row=row, column=2,  value=titulo)
        ws.cell(row=row, column=3,  value=marca)
        ws.cell(row=row, column=4,  value=precio)
        ws.cell(row=row, column=5,  value=link)
        ws.cell(row=row, column=6,  value=img)
        ws.cell(row=row, column=7,  value=stars if stars else "")
        ws.cell(row=row, column=8,  value=resenas if resenas else "")
        ws.cell(row=row, column=9,  value=motor)
        ws.cell(row=row, column=10, value=autonomia)
        ws.cell(row=row, column=11, value=velocidad)
        ws.cell(row=row, column=12, value=rueda)
        ws.cell(row=row, column=13, value=velocidades)
        ws.cell(row=row, column=14, value=peso)
        ws.cell(row=row, column=15, value=item_id)

        for c in range(1, len(COLS) + 1):
            ws.cell(row=row, column=c).border = thin

        row += 1

        if (idx + 1) % 10 == 0:
            print(f"  {idx+1}/{len(todos)} listos...")
        time.sleep(0.2)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Hoja de instrucciones
    ws2 = wb.create_sheet("INSTRUCCIONES")
    instrucciones = [
        ["COMO USAR ESTE ARCHIVO"],
        [],
        ["1. Revisa cada producto en la hoja 'Electricas'"],
        ["2. Columna 'Publicar': SI = aparece en el sitio, NO = se omite"],
        ["3. Puedes corregir el Nombre o Marca si ML los trae mal"],
        ["4. Agrega Motor/Autonomia/Velocidad donde falte (busca en la ficha de ML)"],
        ["5. Guarda el archivo y compartelo — se carga al sitio automaticamente"],
        [],
        ["COLUMNAS CLAVE:"],
        ["  Precio MXN  → precio de venta actual en ML"],
        ["  Link ML     → URL del producto (ya tiene el afiliado si lo agregas)"],
        ["  Imagen URL  → foto principal del producto"],
        ["  Estrellas   → calificacion 0-5 (ej. 4.7)"],
        ["  Resenas     → numero de opiniones"],
    ]
    for r in instrucciones:
        ws2.append(r)
    ws2["A1"].font = Font(bold=True, size=13, color="006847")
    ws2.column_dimensions["A"].width = 70

    out = "resultados_ml.xlsx"
    wb.save(out)
    print(f"\n{'='*60}")
    print(f"  Listo: {out}  ({row-2} productos)")
    print(f"{'='*60}")
    print("\nSiguiente paso:")
    print("  1. Abre resultados_ml.xlsx")
    print("  2. Pon NO en las bicis que no quieras publicar")
    print("  3. Agrega los links de afiliado en la columna Link ML")
    print("     (reemplaza el link normal por el de tu cuenta comparabici)")
    print("  4. Comparte el archivo aqui — lo cargo al sitio en automatico")


if __name__ == "__main__":
    main()
